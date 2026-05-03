import os
import hashlib
import json
import io
from datetime import datetime, date
from functools import wraps
from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, jsonify, send_file, g)
from db import get_db, init_db, log_audit

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'mydsp-secret-key-change-in-production-2026')

# ─── Helpers ────────────────────────────────────────────────────────────────

def hash_password(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('Admin access required.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def current_user():
    return {'id': session.get('user_id'), 'username': session.get('username'),
            'role': session.get('role'), 'full_name': session.get('full_name')}

def log(action, table, record_id=None, old=None, new=None):
    u = current_user()
    log_audit(u['id'], u['username'], action, table, record_id, old, new, request.remote_addr)

# ─── Auth ────────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        db = get_db()
        user = db.execute(
            "SELECT * FROM users WHERE username=%s AND active=1", (username,)
        ).fetchone()
        db.close()
        if user and user['password_hash'] == hash_password(password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            session['full_name'] = user['full_name']
            log('LOGIN', 'users', user['id'])
            return redirect(url_for('dashboard'))
        flash('Invalid username or password.', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    log('LOGOUT', 'users')
    session.clear()
    return redirect(url_for('login'))

# ─── Dashboard ───────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def dashboard():
    db = get_db()
    # Financial KPIs
    total_income = db.execute("SELECT COALESCE(SUM(amount),0) FROM income_entries").fetchone()[0]
    total_expenditure = db.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenditure_entries"
    ).fetchone()[0]
    total_staff_costs = db.execute(
        "SELECT COALESCE(SUM(total_cost),0) FROM staff_costs"
    ).fetchone()[0]
    total_costs = total_expenditure + total_staff_costs
    net_profit = total_income - total_costs

    # Counts
    participant_count = db.execute("SELECT COUNT(*) FROM participants WHERE status='Active'").fetchone()[0]
    open_tasks = db.execute("SELECT COUNT(*) FROM tasks WHERE status NOT IN ('Completed','Closed')").fetchone()[0]
    open_issues = db.execute("SELECT COUNT(*) FROM issues WHERE status='Open'").fetchone()[0]
    petty_cash_balance = db.execute(
        "SELECT COALESCE(SUM(cash_in) - SUM(expense), 0) FROM petty_cash"
    ).fetchone()[0]

    # Recent income (last 10)
    recent_income = db.execute("""
        SELECT entry_date, participant_name, support_category, amount, invoice_number
        FROM income_entries ORDER BY entry_date DESC LIMIT 10
    """).fetchall()

    # Urgent tasks
    urgent_tasks = db.execute("""
        SELECT task_id, title, priority, assigned_to, due_date, status
        FROM tasks WHERE status NOT IN ('Completed','Closed') AND priority LIKE '%%URGENT%%'
        ORDER BY due_date LIMIT 5
    """).fetchall()

    # Critical issues
    critical_issues = db.execute("""
        SELECT issue_id, participant_name, description, risk_level, due_date
        FROM issues WHERE status='Open' AND risk_level IN ('Critical','High')
        ORDER BY due_date LIMIT 5
    """).fetchall()

    # Monthly income chart data (last 6 months)
    monthly_income = db.execute("""
        SELECT month_period, SUM(amount) as total
        FROM income_entries
        WHERE month_period IS NOT NULL
        GROUP BY month_period
        ORDER BY month_period DESC LIMIT 6
    """).fetchall()

    # Income by participant
    income_by_participant = db.execute("""
        SELECT participant_name, SUM(amount) as total
        FROM income_entries
        GROUP BY participant_name
        ORDER BY total DESC LIMIT 8
    """).fetchall()

    db.close()
    return render_template('dashboard.html',
        total_income=total_income, total_costs=total_costs, net_profit=net_profit,
        participant_count=participant_count, open_tasks=open_tasks, open_issues=open_issues,
        petty_cash_balance=petty_cash_balance, recent_income=recent_income,
        urgent_tasks=urgent_tasks, critical_issues=critical_issues,
        monthly_income=monthly_income, income_by_participant=income_by_participant)

# ─── Participants ─────────────────────────────────────────────────────────────

@app.route('/participants')
@login_required
def participants():
    db = get_db()
    rows = db.execute("SELECT * FROM participants ORDER BY full_name").fetchall()
    db.close()
    return render_template('participants.html', participants=rows)

@app.route('/participants/new', methods=['GET', 'POST'])
@login_required
def participant_new():
    if request.method == 'POST':
        f = request.form
        db = get_db()
        # Auto-generate participant_id
        count = db.execute("SELECT COUNT(*) FROM participants").fetchone()[0]
        pid = f"P{count+1:03d}"
        db.execute("""
            INSERT INTO participants (participant_id, full_name, ndis_number, date_of_birth,
            plan_start, plan_end, support_type, primary_diagnosis, plan_manager, total_funding,
            core_funding, cb_funding, sc_funding, iscp_funding, support_coordinator,
            emergency_contact, emergency_phone, language, interpreter_required, status,
            sw_schedule, goals_summary, risk_flag, next_review_date, plan_status, notes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (pid, f['full_name'], f.get('ndis_number'), f.get('date_of_birth'),
              f.get('plan_start'), f.get('plan_end'), f.get('support_type'),
              f.get('primary_diagnosis'), f.get('plan_manager'),
              float(f['total_funding'] or 0), float(f.get('core_funding') or 0),
              float(f.get('cb_funding') or 0), float(f.get('sc_funding') or 0),
              float(f.get('iscp_funding') or 0), f.get('support_coordinator'),
              f.get('emergency_contact'), f.get('emergency_phone'), f.get('language'),
              f.get('interpreter_required'), f.get('status','Active'), f.get('sw_schedule'),
              f.get('goals_summary'), f.get('risk_flag'), f.get('next_review_date'),
              f.get('plan_status'), f.get('notes')))
        new_id = db.execute("SELECT lastval()").fetchone()[0]
        db.commit()
        db.close()
        log('CREATE', 'participants', new_id, new={'name': f['full_name']})
        flash(f'Participant {f["full_name"]} added successfully.', 'success')
        return redirect(url_for('participants'))
    return render_template('participant_form.html', participant=None)

@app.route('/participants/<int:pid>/edit', methods=['GET', 'POST'])
@login_required
def participant_edit(pid):
    db = get_db()
    p = db.execute("SELECT * FROM participants WHERE id=%s", (pid,)).fetchone()
    if not p:
        db.close()
        flash('Participant not found.', 'danger')
        return redirect(url_for('participants'))
    if request.method == 'POST':
        f = request.form
        old = dict(p)
        db.execute("""
            UPDATE participants SET full_name=%s, ndis_number=%s, date_of_birth=%s,
            plan_start=%s, plan_end=%s, support_type=%s, primary_diagnosis=%s, plan_manager=%s,
            total_funding=%s, core_funding=%s, cb_funding=%s, sc_funding=%s, iscp_funding=%s,
            support_coordinator=%s, emergency_contact=%s, emergency_phone=%s, language=%s,
            interpreter_required=%s, status=%s, sw_schedule=%s, goals_summary=%s, risk_flag=%s,
            next_review_date=%s, plan_status=%s, notes=%s, updated_at=NOW()
            WHERE id=%s
        """, (f['full_name'], f.get('ndis_number'), f.get('date_of_birth'),
              f.get('plan_start'), f.get('plan_end'), f.get('support_type'),
              f.get('primary_diagnosis'), f.get('plan_manager'),
              float(f['total_funding'] or 0), float(f.get('core_funding') or 0),
              float(f.get('cb_funding') or 0), float(f.get('sc_funding') or 0),
              float(f.get('iscp_funding') or 0), f.get('support_coordinator'),
              f.get('emergency_contact'), f.get('emergency_phone'), f.get('language'),
              f.get('interpreter_required'), f.get('status','Active'), f.get('sw_schedule'),
              f.get('goals_summary'), f.get('risk_flag'), f.get('next_review_date'),
              f.get('plan_status'), f.get('notes'), pid))
        db.commit()
        db.close()
        log('UPDATE', 'participants', pid, old, {'name': f['full_name']})
        flash('Participant updated.', 'success')
        return redirect(url_for('participants'))
    db.close()
    return render_template('participant_form.html', participant=p)

@app.route('/participants/<int:pid>/delete', methods=['POST'])
@admin_required
def participant_delete(pid):
    db = get_db()
    p = db.execute("SELECT full_name FROM participants WHERE id=%s", (pid,)).fetchone()
    if p:
        db.execute("DELETE FROM participants WHERE id=%s", (pid,))
        db.commit()
        log('DELETE', 'participants', pid, old={'name': p['full_name']})
        flash(f'Participant {p["full_name"]} deleted.', 'success')
    db.close()
    return redirect(url_for('participants'))

# ─── Income Ledger ────────────────────────────────────────────────────────────

@app.route('/income')
@login_required
def income():
    db = get_db()
    month = request.args.get('month', '')
    participant = request.args.get('participant', '')
    query = "SELECT * FROM income_entries WHERE 1=1"
    params = []
    if month:
        query += " AND month_period=%s"
        params.append(month)
    if participant:
        query += " AND participant_name LIKE %s"
        params.append(f'%{participant}%')
    query += " ORDER BY entry_date DESC"
    rows = db.execute(query, params).fetchall()
    total = sum(r['amount'] for r in rows)
    participants_list = db.execute(
        "SELECT id, full_name FROM participants ORDER BY full_name"
    ).fetchall()
    months = db.execute(
        "SELECT DISTINCT month_period FROM income_entries WHERE month_period IS NOT NULL ORDER BY month_period DESC"
    ).fetchall()
    db.close()
    return render_template('income.html', entries=rows, total=total,
                           participants_list=participants_list, months=months,
                           filter_month=month, filter_participant=participant)

@app.route('/income/new', methods=['GET', 'POST'])
@login_required
def income_new():
    if request.method == 'POST':
        f = request.form
        db = get_db()
        db.execute("""
            INSERT INTO income_entries (entry_date, month_period, participant_id, participant_name,
            support_category, ndis_item_code, invoice_number, amount, plan_manager_type, notes, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (f['entry_date'], f.get('month_period'), f.get('participant_id') or None,
              f.get('participant_name'), f.get('support_category'), f.get('ndis_item_code'),
              f.get('invoice_number'), float(f['amount'] or 0), f.get('plan_manager_type'),
              f.get('notes'), session['user_id']))
        new_id = db.execute("SELECT lastval()").fetchone()[0]
        db.commit()
        db.close()
        log('CREATE', 'income_entries', new_id)
        flash('Income entry added.', 'success')
        return redirect(url_for('income'))
    db = get_db()
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    db.close()
    return render_template('income_form.html', entry=None, participants_list=participants_list)

@app.route('/income/<int:eid>/edit', methods=['GET', 'POST'])
@login_required
def income_edit(eid):
    db = get_db()
    entry = db.execute("SELECT * FROM income_entries WHERE id=%s", (eid,)).fetchone()
    if not entry:
        db.close()
        flash('Entry not found.', 'danger')
        return redirect(url_for('income'))
    if request.method == 'POST':
        f = request.form
        old = dict(entry)
        db.execute("""
            UPDATE income_entries SET entry_date=%s, month_period=%s, participant_id=%s,
            participant_name=%s, support_category=%s, ndis_item_code=%s, invoice_number=%s,
            amount=%s, plan_manager_type=%s, notes=%s, updated_at=NOW() WHERE id=%s
        """, (f['entry_date'], f.get('month_period'), f.get('participant_id') or None,
              f.get('participant_name'), f.get('support_category'), f.get('ndis_item_code'),
              f.get('invoice_number'), float(f['amount'] or 0), f.get('plan_manager_type'),
              f.get('notes'), eid))
        db.commit()
        db.close()
        log('UPDATE', 'income_entries', eid, old)
        flash('Income entry updated.', 'success')
        return redirect(url_for('income'))
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    db.close()
    return render_template('income_form.html', entry=entry, participants_list=participants_list)

@app.route('/income/<int:eid>/delete', methods=['POST'])
@login_required
def income_delete(eid):
    db = get_db()
    db.execute("DELETE FROM income_entries WHERE id=%s", (eid,))
    db.commit()
    db.close()
    log('DELETE', 'income_entries', eid)
    flash('Income entry deleted.', 'success')
    return redirect(url_for('income'))

# ─── Expenditure ─────────────────────────────────────────────────────────────

@app.route('/expenditure')
@login_required
def expenditure():
    db = get_db()
    month = request.args.get('month', '')
    cat = request.args.get('category', '')
    query = "SELECT * FROM expenditure_entries WHERE 1=1"
    params = []
    if month:
        query += " AND month_period=%s"
        params.append(month)
    if cat:
        query += " AND category=%s"
        params.append(cat)
    query += " ORDER BY entry_date DESC"
    rows = db.execute(query, params).fetchall()
    total = sum(r['amount'] for r in rows)
    categories = db.execute(
        "SELECT DISTINCT category FROM expenditure_entries WHERE category IS NOT NULL ORDER BY category"
    ).fetchall()
    months = db.execute(
        "SELECT DISTINCT month_period FROM expenditure_entries WHERE month_period IS NOT NULL ORDER BY month_period DESC"
    ).fetchall()
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    db.close()
    return render_template('expenditure.html', entries=rows, total=total,
                           categories=categories, months=months,
                           participants_list=participants_list,
                           filter_month=month, filter_cat=cat)

@app.route('/expenditure/new', methods=['GET', 'POST'])
@login_required
def expenditure_new():
    if request.method == 'POST':
        f = request.form
        db = get_db()
        db.execute("""
            INSERT INTO expenditure_entries (entry_date, month_period, category, sub_category,
            description, supplier, amount, participant_id, participant_name, invoice_number,
            payment_method, notes, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (f['entry_date'], f.get('month_period'), f.get('category'), f.get('sub_category'),
              f.get('description'), f.get('supplier'), float(f['amount'] or 0),
              f.get('participant_id') or None, f.get('participant_name'),
              f.get('invoice_number'), f.get('payment_method'), f.get('notes'), session['user_id']))
        new_id = db.execute("SELECT lastval()").fetchone()[0]
        db.commit()
        db.close()
        log('CREATE', 'expenditure_entries', new_id)
        flash('Expenditure entry added.', 'success')
        return redirect(url_for('expenditure'))
    db = get_db()
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    db.close()
    return render_template('expenditure_form.html', entry=None, participants_list=participants_list)

@app.route('/expenditure/<int:eid>/edit', methods=['GET', 'POST'])
@login_required
def expenditure_edit(eid):
    db = get_db()
    entry = db.execute("SELECT * FROM expenditure_entries WHERE id=%s", (eid,)).fetchone()
    if not entry:
        db.close()
        flash('Entry not found.', 'danger')
        return redirect(url_for('expenditure'))
    if request.method == 'POST':
        f = request.form
        old = dict(entry)
        db.execute("""
            UPDATE expenditure_entries SET entry_date=%s, month_period=%s, category=%s,
            sub_category=%s, description=%s, supplier=%s, amount=%s, participant_id=%s,
            participant_name=%s, invoice_number=%s, payment_method=%s, notes=%s,
            updated_at=NOW() WHERE id=%s
        """, (f['entry_date'], f.get('month_period'), f.get('category'), f.get('sub_category'),
              f.get('description'), f.get('supplier'), float(f['amount'] or 0),
              f.get('participant_id') or None, f.get('participant_name'),
              f.get('invoice_number'), f.get('payment_method'), f.get('notes'), eid))
        db.commit()
        db.close()
        log('UPDATE', 'expenditure_entries', eid, old)
        flash('Expenditure entry updated.', 'success')
        return redirect(url_for('expenditure'))
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    db.close()
    return render_template('expenditure_form.html', entry=entry, participants_list=participants_list)

@app.route('/expenditure/<int:eid>/delete', methods=['POST'])
@login_required
def expenditure_delete(eid):
    db = get_db()
    db.execute("DELETE FROM expenditure_entries WHERE id=%s", (eid,))
    db.commit()
    db.close()
    log('DELETE', 'expenditure_entries', eid)
    flash('Expenditure entry deleted.', 'success')
    return redirect(url_for('expenditure'))

# ─── Staff Costs ──────────────────────────────────────────────────────────────

@app.route('/staff-costs')
@login_required
def staff_costs():
    db = get_db()
    month = request.args.get('month', '')
    staff = request.args.get('staff', '')
    query = "SELECT * FROM staff_costs WHERE 1=1"
    params = []
    if month:
        query += " AND month_period=%s"
        params.append(month)
    if staff:
        query += " AND staff_name LIKE %s"
        params.append(f'%{staff}%')
    query += " ORDER BY pay_date DESC"
    rows = db.execute(query, params).fetchall()
    total_revenue = sum(r['ndis_revenue'] or 0 for r in rows)
    total_cost = sum(r['total_cost'] or 0 for r in rows)
    total_margin = sum(r['margin'] or 0 for r in rows)
    months = db.execute(
        "SELECT DISTINCT month_period FROM staff_costs WHERE month_period IS NOT NULL ORDER BY month_period DESC"
    ).fetchall()
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    ndis_rates = {'Weekday Day': 70.23, 'Weekday Evening': 77.38, 'Saturday': 98.83,
                  'Sunday': 127.43, 'Public Holiday': 156.03, 'SIL Night': 243.56}
    db.close()
    return render_template('staff_costs.html', entries=rows, total_revenue=total_revenue,
                           total_cost=total_cost, total_margin=total_margin, months=months,
                           participants_list=participants_list, ndis_rates=ndis_rates,
                           filter_month=month, filter_staff=staff)

@app.route('/staff-costs/new', methods=['GET', 'POST'])
@login_required
def staff_cost_new():
    if request.method == 'POST':
        f = request.form
        qty = float(f.get('qty_hours') or 0)
        ndis_rate = float(f.get('ndis_rate') or 0)
        actual_rate = float(f.get('actual_rate') or 0)
        ndis_revenue = qty * ndis_rate
        actual_wage = qty * actual_rate
        super_amt = actual_wage * 0.12
        total_cost = actual_wage + super_amt
        margin = ndis_revenue - total_cost
        db = get_db()
        db.execute("""
            INSERT INTO staff_costs (pay_date, month_period, staff_name, role, participant_id,
            participant_name, shift_type, schedule_notes, qty_hours, ndis_rate, ndis_revenue,
            actual_rate, actual_wage, super_amount, total_cost, margin, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (f['pay_date'], f.get('month_period'), f['staff_name'], f.get('role'),
              f.get('participant_id') or None, f.get('participant_name'), f.get('shift_type'),
              f.get('schedule_notes'), qty, ndis_rate, ndis_revenue, actual_rate,
              actual_wage, super_amt, total_cost, margin, session['user_id']))
        new_id = db.execute("SELECT lastval()").fetchone()[0]
        db.commit()
        db.close()
        log('CREATE', 'staff_costs', new_id)
        flash('Staff cost entry added.', 'success')
        return redirect(url_for('staff_costs'))
    db = get_db()
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    ndis_rates = {'Weekday Day': 70.23, 'Weekday Evening': 77.38, 'Saturday': 98.83,
                  'Sunday': 127.43, 'Public Holiday': 156.03, 'SIL Night': 243.56}
    db.close()
    return render_template('staff_cost_form.html', entry=None,
                           participants_list=participants_list, ndis_rates=ndis_rates)

@app.route('/staff-costs/<int:eid>/edit', methods=['GET', 'POST'])
@login_required
def staff_cost_edit(eid):
    db = get_db()
    entry = db.execute("SELECT * FROM staff_costs WHERE id=%s", (eid,)).fetchone()
    if not entry:
        db.close()
        flash('Entry not found.', 'danger')
        return redirect(url_for('staff_costs'))
    if request.method == 'POST':
        f = request.form
        old = dict(entry)
        qty = float(f.get('qty_hours') or 0)
        ndis_rate = float(f.get('ndis_rate') or 0)
        actual_rate = float(f.get('actual_rate') or 0)
        ndis_revenue = qty * ndis_rate
        actual_wage = qty * actual_rate
        super_amt = actual_wage * 0.12
        total_cost = actual_wage + super_amt
        margin = ndis_revenue - total_cost
        db.execute("""
            UPDATE staff_costs SET pay_date=%s, month_period=%s, staff_name=%s, role=%s,
            participant_id=%s, participant_name=%s, shift_type=%s, schedule_notes=%s, qty_hours=%s,
            ndis_rate=%s, ndis_revenue=%s, actual_rate=%s, actual_wage=%s, super_amount=%s,
            total_cost=%s, margin=%s, updated_at=NOW() WHERE id=%s
        """, (f['pay_date'], f.get('month_period'), f['staff_name'], f.get('role'),
              f.get('participant_id') or None, f.get('participant_name'), f.get('shift_type'),
              f.get('schedule_notes'), qty, ndis_rate, ndis_revenue, actual_rate,
              actual_wage, super_amt, total_cost, margin, eid))
        db.commit()
        db.close()
        log('UPDATE', 'staff_costs', eid, old)
        flash('Staff cost entry updated.', 'success')
        return redirect(url_for('staff_costs'))
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    ndis_rates = {'Weekday Day': 70.23, 'Weekday Evening': 77.38, 'Saturday': 98.83,
                  'Sunday': 127.43, 'Public Holiday': 156.03, 'SIL Night': 243.56}
    db.close()
    return render_template('staff_cost_form.html', entry=entry,
                           participants_list=participants_list, ndis_rates=ndis_rates)

@app.route('/staff-costs/<int:eid>/delete', methods=['POST'])
@login_required
def staff_cost_delete(eid):
    db = get_db()
    db.execute("DELETE FROM staff_costs WHERE id=%s", (eid,))
    db.commit()
    db.close()
    log('DELETE', 'staff_costs', eid)
    flash('Staff cost entry deleted.', 'success')
    return redirect(url_for('staff_costs'))

# ─── Petty Cash ───────────────────────────────────────────────────────────────

@app.route('/petty-cash')
@login_required
def petty_cash():
    db = get_db()
    rows = db.execute("SELECT * FROM petty_cash ORDER BY entry_date DESC, id DESC").fetchall()
    balance = db.execute(
        "SELECT COALESCE(SUM(cash_in),0) - COALESCE(SUM(expense),0) FROM petty_cash"
    ).fetchone()[0]
    total_expense = db.execute("SELECT COALESCE(SUM(expense),0) FROM petty_cash").fetchone()[0]
    total_in = db.execute("SELECT COALESCE(SUM(cash_in),0) FROM petty_cash").fetchone()[0]
    db.close()
    return render_template('petty_cash.html', entries=rows, balance=balance,
                           total_expense=total_expense, total_in=total_in)

@app.route('/petty-cash/new', methods=['GET', 'POST'])
@login_required
def petty_cash_new():
    if request.method == 'POST':
        f = request.form
        expense = float(f.get('expense') or 0)
        cash_in = float(f.get('cash_in') or 0)
        db = get_db()
        prev_bal = db.execute(
            "SELECT COALESCE(SUM(cash_in),0) - COALESCE(SUM(expense),0) FROM petty_cash"
        ).fetchone()[0]
        balance = prev_bal + cash_in - expense
        db.execute("""
            INSERT INTO petty_cash (entry_date, description, expense, cash_in, balance,
            location, receipt_obtained, notes, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (f['entry_date'], f['description'], expense, cash_in, balance,
              f.get('location'), f.get('receipt_obtained','No'), f.get('notes'), session['user_id']))
        new_id = db.execute("SELECT lastval()").fetchone()[0]
        db.commit()
        db.close()
        log('CREATE', 'petty_cash', new_id)
        flash('Petty cash entry added.', 'success')
        return redirect(url_for('petty_cash'))
    return render_template('petty_cash_form.html', entry=None)

@app.route('/petty-cash/<int:eid>/delete', methods=['POST'])
@login_required
def petty_cash_delete(eid):
    db = get_db()
    db.execute("DELETE FROM petty_cash WHERE id=%s", (eid,))
    db.commit()
    db.close()
    log('DELETE', 'petty_cash', eid)
    flash('Petty cash entry deleted.', 'success')
    return redirect(url_for('petty_cash'))

# ─── Tasks ────────────────────────────────────────────────────────────────────

@app.route('/tasks')
@login_required
def tasks():
    db = get_db()
    status_filter = request.args.get('status', '')
    priority_filter = request.args.get('priority', '')
    query = "SELECT * FROM tasks WHERE 1=1"
    params = []
    if status_filter:
        query += " AND status=%s"
        params.append(status_filter)
    if priority_filter:
        query += " AND priority LIKE %s"
        params.append(f'%{priority_filter}%')
    query += " ORDER BY CASE priority WHEN 'URGENT' THEN 1 WHEN 'HIGH' THEN 2 WHEN 'MEDIUM' THEN 3 ELSE 4 END, due_date"
    rows = db.execute(query, params).fetchall()
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    db.close()
    return render_template('tasks.html', tasks=rows, participants_list=participants_list,
                           filter_status=status_filter, filter_priority=priority_filter)

@app.route('/tasks/new', methods=['GET', 'POST'])
@login_required
def task_new():
    if request.method == 'POST':
        f = request.form
        db = get_db()
        count = db.execute("SELECT COUNT(*) FROM tasks").fetchone()[0]
        task_id = f"T{count+1:03d}"
        db.execute("""
            INSERT INTO tasks (task_id, title, priority, participant_id, participant_name,
            assigned_to, due_date, status, notes, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (task_id, f['title'], f.get('priority'), f.get('participant_id') or None,
              f.get('participant_name'), f.get('assigned_to'), f.get('due_date'),
              f.get('status','Not Started'), f.get('notes'), session['user_id']))
        new_id = db.execute("SELECT lastval()").fetchone()[0]
        db.commit()
        db.close()
        log('CREATE', 'tasks', new_id)
        flash('Task added.', 'success')
        return redirect(url_for('tasks'))
    db = get_db()
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    db.close()
    return render_template('task_form.html', task=None, participants_list=participants_list)

@app.route('/tasks/<int:tid>/edit', methods=['GET', 'POST'])
@login_required
def task_edit(tid):
    db = get_db()
    task = db.execute("SELECT * FROM tasks WHERE id=%s", (tid,)).fetchone()
    if not task:
        db.close()
        flash('Task not found.', 'danger')
        return redirect(url_for('tasks'))
    if request.method == 'POST':
        f = request.form
        old = dict(task)
        completed = f.get('completed_date') if f.get('status') == 'Completed' else task['completed_date']
        db.execute("""
            UPDATE tasks SET title=%s, priority=%s, participant_id=%s, participant_name=%s,
            assigned_to=%s, due_date=%s, status=%s, completed_date=%s, notes=%s,
            updated_at=NOW() WHERE id=%s
        """, (f['title'], f.get('priority'), f.get('participant_id') or None,
              f.get('participant_name'), f.get('assigned_to'), f.get('due_date'),
              f.get('status'), completed, f.get('notes'), tid))
        db.commit()
        db.close()
        log('UPDATE', 'tasks', tid, old)
        flash('Task updated.', 'success')
        return redirect(url_for('tasks'))
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    db.close()
    return render_template('task_form.html', task=task, participants_list=participants_list)

@app.route('/tasks/<int:tid>/delete', methods=['POST'])
@login_required
def task_delete(tid):
    db = get_db()
    db.execute("DELETE FROM tasks WHERE id=%s", (tid,))
    db.commit()
    db.close()
    log('DELETE', 'tasks', tid)
    flash('Task deleted.', 'success')
    return redirect(url_for('tasks'))

# ─── Communications ───────────────────────────────────────────────────────────

@app.route('/communications')
@login_required
def communications():
    db = get_db()
    participant = request.args.get('participant', '')
    contact_type = request.args.get('contact_type', '')
    query = "SELECT * FROM communications WHERE 1=1"
    params = []
    if participant:
        query += " AND participant_name LIKE %s"
        params.append(f'%{participant}%')
    if contact_type:
        query += " AND contact_type=%s"
        params.append(contact_type)
    query += " ORDER BY comm_date DESC"
    rows = db.execute(query, params).fetchall()
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    db.close()
    return render_template('communications.html', entries=rows,
                           participants_list=participants_list,
                           filter_participant=participant, filter_type=contact_type)

@app.route('/communications/new', methods=['GET', 'POST'])
@login_required
def communication_new():
    if request.method == 'POST':
        f = request.form
        db = get_db()
        db.execute("""
            INSERT INTO communications (comm_date, participant_id, participant_name, contact_type,
            person_assigned, organisation_role, reason, outcome, follow_up, follow_up_date, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (f['comm_date'], f.get('participant_id') or None, f.get('participant_name'),
              f.get('contact_type'), f.get('person_assigned'), f.get('organisation_role'),
              f.get('reason'), f.get('outcome'), f.get('follow_up'), f.get('follow_up_date'),
              session['user_id']))
        new_id = db.execute("SELECT lastval()").fetchone()[0]
        db.commit()
        db.close()
        log('CREATE', 'communications', new_id)
        flash('Communication log entry added.', 'success')
        return redirect(url_for('communications'))
    db = get_db()
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    db.close()
    return render_template('communication_form.html', entry=None, participants_list=participants_list)

@app.route('/communications/<int:eid>/delete', methods=['POST'])
@login_required
def communication_delete(eid):
    db = get_db()
    db.execute("DELETE FROM communications WHERE id=%s", (eid,))
    db.commit()
    db.close()
    log('DELETE', 'communications', eid)
    flash('Communication entry deleted.', 'success')
    return redirect(url_for('communications'))

# ─── Issues ───────────────────────────────────────────────────────────────────

@app.route('/issues')
@login_required
def issues():
    db = get_db()
    status_filter = request.args.get('status', '')
    risk_filter = request.args.get('risk', '')
    query = "SELECT * FROM issues WHERE 1=1"
    params = []
    if status_filter:
        query += " AND status=%s"
        params.append(status_filter)
    if risk_filter:
        query += " AND risk_level=%s"
        params.append(risk_filter)
    query += " ORDER BY CASE risk_level WHEN 'Critical' THEN 1 WHEN 'High' THEN 2 WHEN 'Medium' THEN 3 ELSE 4 END, due_date"
    rows = db.execute(query, params).fetchall()
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    db.close()
    return render_template('issues.html', issues=rows, participants_list=participants_list,
                           filter_status=status_filter, filter_risk=risk_filter)

@app.route('/issues/new', methods=['GET', 'POST'])
@login_required
def issue_new():
    if request.method == 'POST':
        f = request.form
        db = get_db()
        count = db.execute("SELECT COUNT(*) FROM issues").fetchone()[0]
        issue_id = f"I{count+1:03d}"
        db.execute("""
            INSERT INTO issues (issue_id, issue_date, participant_id, participant_name,
            description, risk_level, action_required, assigned_to, due_date, status, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (issue_id, f['issue_date'], f.get('participant_id') or None,
              f.get('participant_name'), f['description'], f.get('risk_level'),
              f.get('action_required'), f.get('assigned_to'), f.get('due_date'),
              f.get('status','Open'), session['user_id']))
        new_id = db.execute("SELECT lastval()").fetchone()[0]
        db.commit()
        db.close()
        log('CREATE', 'issues', new_id)
        flash('Issue logged.', 'success')
        return redirect(url_for('issues'))
    db = get_db()
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    db.close()
    return render_template('issue_form.html', issue=None, participants_list=participants_list)

@app.route('/issues/<int:iid>/edit', methods=['GET', 'POST'])
@login_required
def issue_edit(iid):
    db = get_db()
    issue = db.execute("SELECT * FROM issues WHERE id=%s", (iid,)).fetchone()
    if not issue:
        db.close()
        flash('Issue not found.', 'danger')
        return redirect(url_for('issues'))
    if request.method == 'POST':
        f = request.form
        old = dict(issue)
        db.execute("""
            UPDATE issues SET issue_date=%s, participant_id=%s, participant_name=%s,
            description=%s, risk_level=%s, action_required=%s, assigned_to=%s, due_date=%s,
            status=%s, resolution=%s, updated_at=NOW() WHERE id=%s
        """, (f['issue_date'], f.get('participant_id') or None, f.get('participant_name'),
              f['description'], f.get('risk_level'), f.get('action_required'),
              f.get('assigned_to'), f.get('due_date'), f.get('status'),
              f.get('resolution'), iid))
        db.commit()
        db.close()
        log('UPDATE', 'issues', iid, old)
        flash('Issue updated.', 'success')
        return redirect(url_for('issues'))
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    db.close()
    return render_template('issue_form.html', issue=issue, participants_list=participants_list)

@app.route('/issues/<int:iid>/delete', methods=['POST'])
@login_required
def issue_delete(iid):
    db = get_db()
    db.execute("DELETE FROM issues WHERE id=%s", (iid,))
    db.commit()
    db.close()
    log('DELETE', 'issues', iid)
    flash('Issue deleted.', 'success')
    return redirect(url_for('issues'))

# ─── Reports ──────────────────────────────────────────────────────────────────

@app.route('/reports')
@login_required
def reports():
    db = get_db()
    month = request.args.get('month', '')
    participant = request.args.get('participant', '')

    # Income summary
    income_q = "SELECT participant_name, SUM(amount) as total FROM income_entries WHERE 1=1"
    exp_q = "SELECT category, SUM(amount) as total FROM expenditure_entries WHERE 1=1"
    params = []
    if month:
        income_q += " AND month_period=%s"
        exp_q += " AND month_period=%s"
        params.append(month)
    income_q += " GROUP BY participant_name ORDER BY total DESC"
    exp_q += " GROUP BY category ORDER BY total DESC"
    income_by_participant = db.execute(income_q, params).fetchall()
    exp_by_category = db.execute(exp_q, params).fetchall()

    total_income = sum(r['total'] for r in income_by_participant)
    total_exp = sum(r['total'] for r in exp_by_category)
    total_staff = db.execute(
        "SELECT COALESCE(SUM(total_cost),0) FROM staff_costs" +
        (" WHERE month_period=%s" if month else ""),
        [month] if month else []
    ).fetchone()[0]
    net = total_income - total_exp - total_staff

    months = db.execute(
        "SELECT DISTINCT month_period FROM income_entries WHERE month_period IS NOT NULL ORDER BY month_period DESC"
    ).fetchall()
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()

    staff_summary = db.execute(
        "SELECT staff_name, SUM(qty_hours) as hrs, SUM(ndis_revenue) as rev, SUM(total_cost) as cost, SUM(margin) as margin FROM staff_costs" +
        (" WHERE month_period=%s" if month else "") +
        " GROUP BY staff_name ORDER BY cost DESC",
        [month] if month else []
    ).fetchall()

    db.close()
    return render_template('reports.html',
        income_by_participant=income_by_participant,
        exp_by_category=exp_by_category, staff_summary=staff_summary,
        total_income=total_income, total_exp=total_exp, total_staff=total_staff,
        net=net, months=months, participants_list=participants_list,
        filter_month=month, filter_participant=participant)

@app.route('/reports/export/excel')
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db = get_db()
    month = request.args.get('month', '')
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='1a3c5e')
    alt_fill = PatternFill('solid', start_color='EBF3FB')

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # Income sheet
    ws1 = wb.active
    ws1.title = 'Income Ledger'
    hdrs = ['Date', 'Month', 'Participant', 'Support Category', 'NDIS Item Code',
            'Invoice No.', 'Amount ($)', 'Plan Manager', 'Notes']
    style_header(ws1, hdrs)
    q = "SELECT entry_date, month_period, participant_name, support_category, ndis_item_code, invoice_number, amount, plan_manager_type, notes FROM income_entries"
    params = []
    if month:
        q += " WHERE month_period=%s"
        params.append(month)
    q += " ORDER BY entry_date DESC"
    for r_idx, row in enumerate(db.execute(q, params).fetchall(), 2):
        for c_idx, val in enumerate(row, 1):
            ws1.cell(row=r_idx, column=c_idx, value=val)
        if r_idx % 2 == 0:
            for c in range(1, len(hdrs)+1):
                ws1.cell(row=r_idx, column=c).fill = alt_fill
    auto_width(ws1)

    # Expenditure sheet
    ws2 = wb.create_sheet('Expenditure')
    hdrs2 = ['Date', 'Month', 'Category', 'Sub-Category', 'Description',
             'Supplier', 'Amount ($)', 'Participant', 'Payment Method', 'Notes']
    style_header(ws2, hdrs2)
    q2 = "SELECT entry_date, month_period, category, sub_category, description, supplier, amount, participant_name, payment_method, notes FROM expenditure_entries"
    params2 = []
    if month:
        q2 += " WHERE month_period=%s"
        params2.append(month)
    q2 += " ORDER BY entry_date DESC"
    for r_idx, row in enumerate(db.execute(q2, params2).fetchall(), 2):
        for c_idx, val in enumerate(row, 1):
            ws2.cell(row=r_idx, column=c_idx, value=val)
        if r_idx % 2 == 0:
            for c in range(1, len(hdrs2)+1):
                ws2.cell(row=r_idx, column=c).fill = alt_fill
    auto_width(ws2)

    # Staff Costs sheet
    ws3 = wb.create_sheet('Staff Costs')
    hdrs3 = ['Pay Date', 'Month', 'Staff Name', 'Role', 'Participant', 'Shift Type',
             'Qty (Hrs)', 'NDIS Rate', 'NDIS Revenue ($)', 'Actual Rate',
             'Actual Wage ($)', 'Super ($)', 'Total Cost ($)', 'Margin ($)']
    style_header(ws3, hdrs3)
    q3 = "SELECT pay_date, month_period, staff_name, role, participant_name, shift_type, qty_hours, ndis_rate, ndis_revenue, actual_rate, actual_wage, super_amount, total_cost, margin FROM staff_costs"
    params3 = []
    if month:
        q3 += " WHERE month_period=%s"
        params3.append(month)
    q3 += " ORDER BY pay_date DESC"
    for r_idx, row in enumerate(db.execute(q3, params3).fetchall(), 2):
        for c_idx, val in enumerate(row, 1):
            ws3.cell(row=r_idx, column=c_idx, value=val)
        if r_idx % 2 == 0:
            for c in range(1, len(hdrs3)+1):
                ws3.cell(row=r_idx, column=c).fill = alt_fill
    auto_width(ws3)

    # Participants sheet
    ws4 = wb.create_sheet('Participants')
    hdrs4 = ['ID', 'Name', 'NDIS Number', 'DOB', 'Plan Start', 'Plan End',
             'Support Type', 'Diagnosis', 'Plan Manager', 'Total Funding ($)',
             'Core ($)', 'CB ($)', 'SC ($)', 'Status', 'Risk Flag']
    style_header(ws4, hdrs4)
    for r_idx, row in enumerate(db.execute("""
        SELECT participant_id, full_name, ndis_number, date_of_birth, plan_start, plan_end,
        support_type, primary_diagnosis, plan_manager, total_funding, core_funding, cb_funding,
        sc_funding, status, risk_flag FROM participants ORDER BY full_name
    """).fetchall(), 2):
        for c_idx, val in enumerate(row, 1):
            ws4.cell(row=r_idx, column=c_idx, value=val)
        if r_idx % 2 == 0:
            for c in range(1, len(hdrs4)+1):
                ws4.cell(row=r_idx, column=c).fill = alt_fill
    auto_width(ws4)

    db.close()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"MY_DSP_Report_{month or 'All'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/reports/export/pdf')
@login_required
def export_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    db = get_db()
    month = request.args.get('month', '')
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=16,
                                  textColor=colors.HexColor('#1a3c5e'))
    heading_style = ParagraphStyle('heading', parent=styles['Heading2'],
                                    textColor=colors.HexColor('#1a3c5e'), fontSize=12)
    navy = colors.HexColor('#1a3c5e')
    light_blue = colors.HexColor('#EBF3FB')

    def make_table(data, col_widths=None):
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), navy),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_blue]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 4),
        ]))
        return t

    story.append(Paragraph(f'MY DSP — Financial Report {month or "(All Periods)"}', title_style))
    story.append(Paragraph(f'Generated: {datetime.now().strftime("%d %B %Y %H:%M")}', styles['Normal']))
    story.append(Spacer(1, 0.5*cm))

    # Summary
    total_income = db.execute("SELECT COALESCE(SUM(amount),0) FROM income_entries" +
                               (" WHERE month_period=%s" if month else ""),
                               [month] if month else []).fetchone()[0]
    total_exp = db.execute("SELECT COALESCE(SUM(amount),0) FROM expenditure_entries" +
                            (" WHERE month_period=%s" if month else ""),
                            [month] if month else []).fetchone()[0]
    total_staff = db.execute("SELECT COALESCE(SUM(total_cost),0) FROM staff_costs" +
                              (" WHERE month_period=%s" if month else ""),
                              [month] if month else []).fetchone()[0]
    net = total_income - total_exp - total_staff

    summary_data = [
        ['Metric', 'Amount ($)'],
        ['Total Income', f'{total_income:,.2f}'],
        ['Total Expenditure', f'{total_exp:,.2f}'],
        ['Total Staff Costs', f'{total_staff:,.2f}'],
        ['Net Profit / (Loss)', f'{net:,.2f}'],
    ]
    story.append(Paragraph('Financial Summary', heading_style))
    story.append(make_table(summary_data, [10*cm, 5*cm]))
    story.append(Spacer(1, 0.5*cm))

    # Income by Participant
    story.append(Paragraph('Income by Participant', heading_style))
    inc_data = [['Participant', 'Total Income ($)']]
    for row in db.execute("SELECT participant_name, SUM(amount) as total FROM income_entries" +
                           (" WHERE month_period=%s" if month else "") +
                           " GROUP BY participant_name ORDER BY total DESC",
                           [month] if month else []).fetchall():
        inc_data.append([row[0] or '-', f'{row[1]:,.2f}'])
    story.append(make_table(inc_data, [14*cm, 5*cm]))
    story.append(Spacer(1, 0.5*cm))

    # Income Ledger
    story.append(Paragraph('Income Ledger', heading_style))
    inc_ledger = [['Date', 'Participant', 'Support Category', 'Invoice No.', 'Amount ($)']]
    for row in db.execute("SELECT entry_date, participant_name, support_category, invoice_number, amount FROM income_entries" +
                           (" WHERE month_period=%s" if month else "") +
                           " ORDER BY entry_date DESC LIMIT 50",
                           [month] if month else []).fetchall():
        inc_ledger.append([row[0] or '', row[1] or '', row[2] or '', row[3] or '', f'{row[4]:,.2f}'])
    story.append(make_table(inc_ledger, [3*cm, 5*cm, 6*cm, 4*cm, 3*cm]))

    db.close()
    doc.build(story)
    buf.seek(0)
    fname = f"MY_DSP_Report_{month or 'All'}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/pdf')

# ─── Admin ────────────────────────────────────────────────────────────────────

@app.route('/admin/users')
@admin_required
def admin_users():
    db = get_db()
    users = db.execute("SELECT * FROM users ORDER BY username").fetchall()
    db.close()
    return render_template('admin_users.html', users=users)

@app.route('/admin/users/new', methods=['POST'])
@admin_required
def admin_user_new():
    f = request.form
    if not f.get('username') or not f.get('password'):
        flash('Username and password are required.', 'danger')
        return redirect(url_for('admin_users'))
    db = get_db()
    try:
        db.execute("""
            INSERT INTO users (username, password_hash, full_name, role)
            VALUES (%s, %s, %s, %s)
        """, (f['username'].strip(), hash_password(f['password']),
              f.get('full_name', ''), f.get('role', 'staff')))
        db.commit()
        flash(f'User {f["username"]} created.', 'success')
    except Exception as e:
        flash(f'Error: {e}', 'danger')
    db.close()
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:uid>/toggle', methods=['POST'])
@admin_required
def admin_user_toggle(uid):
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id=%s", (uid,)).fetchone()
    if user and user['username'] != 'admin':
        new_active = 0 if user['active'] else 1
        db.execute("UPDATE users SET active=%s WHERE id=%s", (new_active, uid))
        db.commit()
        flash(f'User {user["username"]} {"activated" if new_active else "deactivated"}.', 'success')
    db.close()
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:uid>/reset-password', methods=['POST'])
@admin_required
def admin_reset_password(uid):
    new_pw = request.form.get('new_password', '')
    if len(new_pw) < 6:
        flash('Password must be at least 6 characters.', 'danger')
        return redirect(url_for('admin_users'))
    db = get_db()
    db.execute("UPDATE users SET password_hash=%s WHERE id=%s", (hash_password(new_pw), uid))
    db.commit()
    db.close()
    flash('Password reset successfully.', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/audit-log')
@admin_required
def audit_log():
    db = get_db()
    logs = db.execute("""
        SELECT * FROM audit_log ORDER BY created_at DESC LIMIT 200
    """).fetchall()
    db.close()
    return render_template('audit_log.html', logs=logs)

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current = request.form.get('current_password', '')
        new_pw = request.form.get('new_password', '')
        confirm = request.form.get('confirm_password', '')
        if new_pw != confirm:
            flash('New passwords do not match.', 'danger')
            return redirect(url_for('change_password'))
        if len(new_pw) < 6:
            flash('Password must be at least 6 characters.', 'danger')
            return redirect(url_for('change_password'))
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE id=%s", (session['user_id'],)).fetchone()
        if user['password_hash'] != hash_password(current):
            db.close()
            flash('Current password is incorrect.', 'danger')
            return redirect(url_for('change_password'))
        db.execute("UPDATE users SET password_hash=%s WHERE id=%s",
                   (hash_password(new_pw), session['user_id']))
        db.commit()
        db.close()
        flash('Password changed successfully.', 'success')
        return redirect(url_for('dashboard'))
    return render_template('change_password.html')

# ─── API endpoints for AJAX ───────────────────────────────────────────────────

@app.route('/api/participants')
@login_required
def api_participants():
    db = get_db()
    rows = db.execute("SELECT id, full_name, participant_id FROM participants ORDER BY full_name").fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/ndis-rates')
@login_required
def api_ndis_rates():
    return jsonify({
        'Weekday Day': 70.23, 'Weekday Evening': 77.38, 'Saturday': 98.83,
        'Sunday': 127.43, 'Public Holiday': 156.03, 'SIL Night': 243.56,
        'SIL Active Night': 38.56
    })

# ─── Alerts API ───────────────────────────────────────────────────────────────

@app.route('/api/alerts')
@login_required
def api_alerts():
    from datetime import date, timedelta
    db = get_db()
    today = date.today()
    alerts = []

    # Plans ending within 60 days
    cutoff_60 = (today + timedelta(days=60)).isoformat()
    cutoff_30 = (today + timedelta(days=30)).isoformat()
    cutoff_14 = (today + timedelta(days=14)).isoformat()
    today_str = today.isoformat()

    plans = db.execute("""
        SELECT full_name, plan_end, participant_id FROM participants
        WHERE status='Active' AND plan_end IS NOT NULL AND plan_end != ''
        AND plan_end <= %s ORDER BY plan_end
    """, (cutoff_60,)).fetchall()
    for p in plans:
        end = p['plan_end'][:10] if p['plan_end'] else ''
        if end <= today_str:
            level = 'danger'
            msg = f"EXPIRED: {p['full_name']}'s NDIS plan expired on {end}"
        elif end <= cutoff_14:
            level = 'danger'
            msg = f"URGENT: {p['full_name']}'s plan ends in ≤14 days ({end})"
        elif end <= cutoff_30:
            level = 'warning'
            msg = f"WARNING: {p['full_name']}'s plan ends in ≤30 days ({end})"
        else:
            level = 'info'
            msg = f"NOTICE: {p['full_name']}'s plan ends in ≤60 days ({end})"
        alerts.append({'type': 'plan', 'level': level, 'message': msg})

    # Tasks overdue or due within 3 days
    cutoff_3 = (today + timedelta(days=3)).isoformat()
    tasks = db.execute("""
        SELECT task_id, title, due_date, assigned_to FROM tasks
        WHERE status NOT IN ('Completed','Closed')
        AND due_date IS NOT NULL AND due_date != '' AND due_date <= %s
        ORDER BY due_date
    """, (cutoff_3,)).fetchall()
    for t in tasks:
        due = t['due_date'][:10] if t['due_date'] else ''
        if due < today_str:
            level = 'danger'
            msg = f"OVERDUE TASK [{t['task_id']}]: {t['title'][:60]} — was due {due}"
        else:
            level = 'warning'
            msg = f"TASK DUE SOON [{t['task_id']}]: {t['title'][:60]} — due {due}"
        alerts.append({'type': 'task', 'level': level, 'message': msg})

    # Budget exceeded warnings
    budget_rows = db.execute("""
        SELECT b.month_period, b.category, b.participant_name, b.budget_amount, b.description
        FROM budgets b WHERE b.budget_amount > 0
    """).fetchall()
    for b in budget_rows:
        actual = 0
        if b['category'] == 'Income':
            row = db.execute("""
                SELECT COALESCE(SUM(amount),0) as total FROM income_entries
                WHERE month_period=%s AND (participant_name=%s OR %s='')
            """, (b['month_period'], b['participant_name'] or '', b['participant_name'] or '')).fetchone()
            actual = row['total'] if row else 0
            if actual > b['budget_amount'] * 1.05:
                alerts.append({'type': 'budget', 'level': 'info',
                    'message': f"Income OVER budget for {b['participant_name'] or 'All'} ({b['month_period']}): Actual ${actual:,.0f} vs Budget ${b['budget_amount']:,.0f}"})
        elif b['category'] == 'Staff Costs':
            row = db.execute("""
                SELECT COALESCE(SUM(total_cost),0) as total FROM staff_costs WHERE month_period=%s
            """, (b['month_period'],)).fetchone()
            actual = row['total'] if row else 0
            if actual > b['budget_amount']:
                alerts.append({'type': 'budget', 'level': 'warning',
                    'message': f"Staff costs EXCEED budget ({b['month_period']}): Actual ${actual:,.0f} vs Budget ${b['budget_amount']:,.0f} — over by ${actual - b['budget_amount']:,.0f}"})
        elif b['category'] == 'Expenditure':
            row = db.execute("""
                SELECT COALESCE(SUM(amount),0) as total FROM expenditure_entries WHERE month_period=%s
            """, (b['month_period'],)).fetchone()
            actual = row['total'] if row else 0
            if actual > b['budget_amount']:
                alerts.append({'type': 'budget', 'level': 'warning',
                    'message': f"Expenditure EXCEEDS budget ({b['month_period']}): Actual ${actual:,.0f} vs Budget ${b['budget_amount']:,.0f}"})

    # Staff cost margin warnings (negative margin)
    neg_margin = db.execute("""
        SELECT staff_name, SUM(margin) as total_margin, SUM(total_cost) as total_cost
        FROM staff_costs GROUP BY staff_name HAVING SUM(margin) < 0
    """).fetchall()
    for s in neg_margin:
        alerts.append({'type': 'staff', 'level': 'warning',
            'message': f"Support worker cost alert: {s['staff_name']} has negative margin (costs ${s['total_cost']:,.0f}, margin ${s['total_margin']:,.0f})"})

    db.close()
    return jsonify({'alerts': alerts, 'count': len(alerts)})

# ─── Budget ───────────────────────────────────────────────────────────────────

@app.route('/budget')
@login_required
def budget():
    db = get_db()
    month = request.args.get('month', '')

    q = "SELECT * FROM budgets WHERE 1=1"
    params = []
    if month:
        q += " AND month_period=%s"
        params.append(month)
    q += " ORDER BY month_period DESC, category, description"
    budget_rows = db.execute(q, params).fetchall()

    months_list = db.execute(
        "SELECT DISTINCT month_period FROM budgets ORDER BY month_period DESC"
    ).fetchall()

    # Build comparison: budget vs actual per category/month
    comparison = []
    periods = [month] if month else [r['month_period'] for r in db.execute(
        "SELECT DISTINCT month_period FROM budgets ORDER BY month_period DESC LIMIT 3"
    ).fetchall()]

    for period in periods:
        # Income
        income_budget = db.execute(
            "SELECT COALESCE(SUM(budget_amount),0) FROM budgets WHERE month_period=%s AND category='Income'",
            (period,)).fetchone()[0]
        income_actual = db.execute(
            "SELECT COALESCE(SUM(amount),0) FROM income_entries WHERE month_period=%s",
            (period,)).fetchone()[0]
        # Expenditure
        exp_budget = db.execute(
            "SELECT COALESCE(SUM(budget_amount),0) FROM budgets WHERE month_period=%s AND category='Expenditure'",
            (period,)).fetchone()[0]
        exp_actual = db.execute(
            "SELECT COALESCE(SUM(amount),0) FROM expenditure_entries WHERE month_period=%s",
            (period,)).fetchone()[0]
        # Staff
        staff_budget = db.execute(
            "SELECT COALESCE(SUM(budget_amount),0) FROM budgets WHERE month_period=%s AND category='Staff Costs'",
            (period,)).fetchone()[0]
        staff_actual = db.execute(
            "SELECT COALESCE(SUM(total_cost),0) FROM staff_costs WHERE month_period=%s",
            (period,)).fetchone()[0]

        net_budget = income_budget - exp_budget - staff_budget
        net_actual = income_actual - exp_actual - staff_actual
        comparison.append({
            'period': period,
            'income_budget': income_budget, 'income_actual': income_actual,
            'income_var': income_actual - income_budget,
            'exp_budget': exp_budget, 'exp_actual': exp_actual,
            'exp_var': exp_actual - exp_budget,
            'staff_budget': staff_budget, 'staff_actual': staff_actual,
            'staff_var': staff_actual - staff_budget,
            'net_budget': net_budget, 'net_actual': net_actual,
            'net_var': net_actual - net_budget,
        })

    # Per-participant income budget vs actual
    participant_budgets = db.execute("""
        SELECT b.participant_name, b.month_period, b.budget_amount,
               COALESCE((SELECT SUM(i.amount) FROM income_entries i
                         WHERE i.month_period=b.month_period
                         AND i.participant_name=b.participant_name),0) as actual
        FROM budgets b WHERE b.category='Income' AND b.participant_name IS NOT NULL
        AND b.participant_name != ''
        """ + (" AND b.month_period=%s" if month else "") + """
        ORDER BY b.month_period DESC, b.participant_name
    """, [month] if month else []).fetchall()

    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    db.close()
    return render_template('budget.html', budget_rows=budget_rows, months_list=months_list,
                           comparison=comparison, participant_budgets=participant_budgets,
                           participants_list=participants_list, filter_month=month)

@app.route('/budget/new', methods=['GET', 'POST'])
@login_required
def budget_new():
    if request.method == 'POST':
        f = request.form
        db = get_db()
        db.execute("""
            INSERT INTO budgets (month_period, category, sub_category, description,
            participant_name, budget_amount, notes, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """, (f['month_period'], f['category'], f.get('sub_category'),
              f.get('description'), f.get('participant_name'),
              float(f['budget_amount'] or 0), f.get('notes'), session['user_id']))
        db.commit()
        db.close()
        flash('Budget entry added.', 'success')
        return redirect(url_for('budget'))
    db = get_db()
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    months_list = db.execute(
        "SELECT DISTINCT month_period FROM income_entries ORDER BY month_period DESC"
    ).fetchall()
    db.close()
    return render_template('budget_form.html', entry=None,
                           participants_list=participants_list, months_list=months_list)

@app.route('/budget/<int:bid>/edit', methods=['GET', 'POST'])
@login_required
def budget_edit(bid):
    db = get_db()
    entry = db.execute("SELECT * FROM budgets WHERE id=%s", (bid,)).fetchone()
    if not entry:
        db.close()
        flash('Budget entry not found.', 'danger')
        return redirect(url_for('budget'))
    if request.method == 'POST':
        f = request.form
        db.execute("""
            UPDATE budgets SET month_period=%s, category=%s, sub_category=%s, description=%s,
            participant_name=%s, budget_amount=%s, notes=%s, updated_at=NOW()
            WHERE id=%s
        """, (f['month_period'], f['category'], f.get('sub_category'),
              f.get('description'), f.get('participant_name'),
              float(f['budget_amount'] or 0), f.get('notes'), bid))
        db.commit()
        db.close()
        flash('Budget entry updated.', 'success')
        return redirect(url_for('budget'))
    participants_list = db.execute("SELECT id, full_name FROM participants ORDER BY full_name").fetchall()
    months_list = db.execute(
        "SELECT DISTINCT month_period FROM income_entries ORDER BY month_period DESC"
    ).fetchall()
    db.close()
    return render_template('budget_form.html', entry=entry,
                           participants_list=participants_list, months_list=months_list)

@app.route('/budget/<int:bid>/delete', methods=['POST'])
@login_required
def budget_delete(bid):
    db = get_db()
    db.execute("DELETE FROM budgets WHERE id=%s", (bid,))
    db.commit()
    db.close()
    flash('Budget entry deleted.', 'success')
    return redirect(url_for('budget'))

# ─── Petty Cash Reconciliation ────────────────────────────────────────────────

@app.route('/petty-cash/reconciliation')
@login_required
def petty_cash_recon():
    db = get_db()
    month = request.args.get('month', '')
    months = db.execute(
        """SELECT DISTINCT LEFT(entry_date, 7) as m FROM petty_cash
           WHERE entry_date IS NOT NULL ORDER BY m DESC"""
    ).fetchall()

    entries = []
    opening_balance = 0.0
    if month:
        settings = db.execute(
            "SELECT * FROM petty_cash_settings WHERE month_period=%s", (month,)
        ).fetchone()
        if settings:
            opening_balance = settings['opening_balance'] or 0.0

        entries = db.execute("""
            SELECT * FROM petty_cash
            WHERE LEFT(entry_date, 7) = %s
            ORDER BY entry_date, id
        """, (month,)).fetchall()

    total_in = sum(r['cash_in'] or 0 for r in entries)
    total_out = sum(r['expense'] or 0 for r in entries)
    closing_balance = opening_balance + total_in - total_out

    # Category breakdown
    categories = {}
    for r in entries:
        if r['expense'] and r['expense'] > 0:
            loc = r['location'] or 'General'
            categories[loc] = categories.get(loc, 0) + r['expense']

    db.close()
    return render_template('petty_cash_recon.html',
        month=month, months=months, entries=entries,
        opening_balance=opening_balance, total_in=total_in,
        total_out=total_out, closing_balance=closing_balance,
        categories=categories)

@app.route('/petty-cash/reconciliation/save-settings', methods=['POST'])
@login_required
def petty_cash_recon_settings():
    f = request.form
    db = get_db()
    db.execute("""
        INSERT INTO petty_cash_settings (month_period, opening_balance, monthly_topup, notes)
        VALUES (%s,%s,%s,%s)
        ON CONFLICT(month_period) DO UPDATE SET
        opening_balance=excluded.opening_balance,
        monthly_topup=excluded.monthly_topup,
        notes=excluded.notes,
        updated_at=NOW()
    """, (f['month_period'], float(f.get('opening_balance') or 0),
          float(f.get('monthly_topup') or 500), f.get('notes')))
    db.commit()
    db.close()
    flash('Petty cash settings saved.', 'success')
    return redirect(url_for('petty_cash_recon', month=f['month_period']))

@app.route('/petty-cash/reconciliation/export/pdf')
@login_required
def petty_cash_recon_pdf():
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    db = get_db()
    month = request.args.get('month', '')
    settings = db.execute(
        "SELECT * FROM petty_cash_settings WHERE month_period=%s", (month,)
    ).fetchone()
    opening_balance = settings['opening_balance'] if settings else 0.0

    entries = db.execute("""
        SELECT * FROM petty_cash WHERE LEFT(entry_date, 7) = %s
        ORDER BY entry_date, id
    """, (month,)).fetchall() if month else []

    total_in = sum(r['cash_in'] or 0 for r in entries)
    total_out = sum(r['expense'] or 0 for r in entries)
    closing_balance = opening_balance + total_in - total_out
    db.close()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm,
                             leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()
    navy = colors.HexColor('#1a3c5e')
    light = colors.HexColor('#EBF3FB')
    story = []

    title_style = ParagraphStyle('t', parent=styles['Title'], fontSize=16, textColor=navy, spaceAfter=4)
    sub_style = ParagraphStyle('s', parent=styles['Normal'], fontSize=10, textColor=colors.grey, spaceAfter=12)
    head_style = ParagraphStyle('h', parent=styles['Heading2'], fontSize=12, textColor=navy, spaceBefore=14, spaceAfter=6)

    story.append(Paragraph('MY DSP — Petty Cash Reconciliation', title_style))
    story.append(Paragraph(f'Period: {month}   |   Prepared: {datetime.now().strftime("%d %B %Y")}', sub_style))
    story.append(HRFlowable(width='100%', thickness=2, color=navy))
    story.append(Spacer(1, 0.4*cm))

    # Summary table
    summary = [
        ['', ''],
        ['Opening Balance', f'${opening_balance:,.2f}'],
        ['Add: Cash Received (Top-ups)', f'${total_in:,.2f}'],
        ['Less: Total Expenditure', f'(${total_out:,.2f})'],
        ['', ''],
        ['Closing Balance', f'${closing_balance:,.2f}'],
    ]
    ts = Table(summary, colWidths=[10*cm, 5*cm])
    ts.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 11),
        ('FONTNAME', (0,5), (-1,5), 'Helvetica-Bold'),
        ('FONTSIZE', (0,5), (-1,5), 13),
        ('TEXTCOLOR', (0,5), (-1,5), navy),
        ('LINEABOVE', (0,5), (-1,5), 1.5, navy),
        ('LINEBELOW', (0,5), (-1,5), 1.5, navy),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(ts)
    story.append(Spacer(1, 0.6*cm))

    # Transaction detail
    story.append(Paragraph('Transaction Detail', head_style))
    t_data = [['Date', 'Description', 'Location', 'Expense ($)', 'Cash In ($)', 'Balance ($)', 'Receipt']]
    running = opening_balance
    for r in entries:
        running = running + (r['cash_in'] or 0) - (r['expense'] or 0)
        t_data.append([
            r['entry_date'] or '', r['description'] or '', r['location'] or '',
            f"${r['expense']:,.2f}" if r['expense'] else '-',
            f"${r['cash_in']:,.2f}" if r['cash_in'] else '-',
            f"${running:,.2f}",
            r['receipt_obtained'] or 'No'
        ])
    t_data.append(['', 'TOTALS', '', f'${total_out:,.2f}', f'${total_in:,.2f}', f'${closing_balance:,.2f}', ''])

    tbl = Table(t_data, colWidths=[2.2*cm, 5.5*cm, 2.5*cm, 2.2*cm, 2.2*cm, 2.2*cm, 1.7*cm], repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), navy),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, light]),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#dce8f5')),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ALIGN', (3,0), (-2,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 3),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 1*cm))
    story.append(Paragraph('Authorised by: ________________________    Date: ________________________', styles['Normal']))

    doc.build(story)
    buf.seek(0)
    fname = f"PettyCash_Recon_{month}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/pdf')

@app.route('/petty-cash/reconciliation/export/excel')
@login_required
def petty_cash_recon_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db = get_db()
    month = request.args.get('month', '')
    settings = db.execute(
        "SELECT * FROM petty_cash_settings WHERE month_period=%s", (month,)
    ).fetchone()
    opening_balance = settings['opening_balance'] if settings else 0.0

    entries = db.execute("""
        SELECT * FROM petty_cash WHERE LEFT(entry_date, 7) = %s
        ORDER BY entry_date, id
    """, (month,)).fetchall() if month else []
    db.close()

    total_in = sum(r['cash_in'] or 0 for r in entries)
    total_out = sum(r['expense'] or 0 for r in entries)
    closing_balance = opening_balance + total_in - total_out

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Petty Cash Recon {month}'

    navy_fill = PatternFill('solid', start_color='1a3c5e')
    light_fill = PatternFill('solid', start_color='EBF3FB')
    white_font = Font(color='FFFFFF', bold=True, name='Arial', size=11)
    navy_font = Font(color='1a3c5e', bold=True, name='Arial', size=11)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)

    def hdr(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font = white_font; c.fill = navy_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        return c

    def cell(ws, row, col, val, bold=False, fmt=None, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = bold_font if bold else normal_font
        if fmt: c.number_format = fmt
        if fill: c.fill = fill
        return c

    # Title
    ws.merge_cells('A1:G1')
    t = ws['A1']
    t.value = f'MY DSP — Petty Cash Reconciliation   Period: {month}'
    t.font = Font(name='Arial', size=14, bold=True, color='1a3c5e')
    t.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:G2')
    ws['A2'].value = f'Prepared: {datetime.now().strftime("%d %B %Y")}'
    ws['A2'].font = Font(name='Arial', size=10, color='888888', italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Summary
    row = 4
    cell(ws, row, 1, 'Opening Balance', bold=True)
    cell(ws, row, 2, opening_balance, fmt='$#,##0.00')
    row += 1
    cell(ws, row, 1, 'Add: Cash Received')
    cell(ws, row, 2, total_in, fmt='$#,##0.00')
    row += 1
    cell(ws, row, 1, 'Less: Total Expenditure')
    cell(ws, row, 2, -total_out, fmt='$#,##0.00')
    row += 1
    c = cell(ws, row, 1, 'Closing Balance', bold=True)
    c.font = Font(name='Arial', size=12, bold=True, color='1a3c5e')
    v = cell(ws, row, 2, closing_balance, bold=True, fmt='$#,##0.00')
    v.font = Font(name='Arial', size=12, bold=True, color='1a3c5e')

    # Table
    row += 2
    for col, h in enumerate(['Date', 'Description', 'Location', 'Expense ($)', 'Cash In ($)', 'Running Balance ($)', 'Receipt%s'], 1):
        hdr(ws, row, col, h)
    ws.row_dimensions[row].height = 20

    running = opening_balance
    for i, e in enumerate(entries):
        row += 1
        running = running + (e['cash_in'] or 0) - (e['expense'] or 0)
        fill = light_fill if i % 2 == 0 else None
        cell(ws, row, 1, e['entry_date'] or '', fill=fill)
        cell(ws, row, 2, e['description'] or '', fill=fill)
        cell(ws, row, 3, e['location'] or '', fill=fill)
        cell(ws, row, 4, e['expense'] or 0, fmt='$#,##0.00', fill=fill)
        cell(ws, row, 5, e['cash_in'] or 0, fmt='$#,##0.00', fill=fill)
        cell(ws, row, 6, running, fmt='$#,##0.00', fill=fill)
        cell(ws, row, 7, e['receipt_obtained'] or 'No', fill=fill)

    # Totals row
    row += 1
    tot_fill = PatternFill('solid', start_color='dce8f5')
    cell(ws, row, 1, '', fill=tot_fill)
    cell(ws, row, 2, 'TOTALS', bold=True, fill=tot_fill)
    cell(ws, row, 3, '', fill=tot_fill)
    cell(ws, row, 4, total_out, bold=True, fmt='$#,##0.00', fill=tot_fill)
    cell(ws, row, 5, total_in, bold=True, fmt='$#,##0.00', fill=tot_fill)
    cell(ws, row, 6, closing_balance, bold=True, fmt='$#,##0.00', fill=tot_fill)

    # Sign-off
    row += 2
    ws.cell(row=row, column=1, value='Authorised by: _________________________').font = Font(name='Arial', size=10)
    ws.cell(row=row, column=5, value='Date: _________________________').font = Font(name='Arial', size=10)

    # Column widths
    for col, w in zip('ABCDEFG', [12, 35, 16, 14, 14, 20, 10]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"PettyCash_Recon_{month}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.context_processor
def inject_now():
    return {'now': datetime.now()}

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug)
